from netmiko import ConnectHandler
import openpyxl
import os
import time

EXCEL_FILE = "tabel1.xlsx"
SSH_USER = "admin"
SSH_PASS = ""

def main():
    if not os.path.exists(EXCEL_FILE): 
        return
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    network_data = []
    for row_idx in range(4, sheet.max_row + 1):
        router_name = sheet.cell(row=row_idx, column=2).value
        ip_ssh = sheet.cell(row=row_idx, column=3).value
        vlan_name = sheet.cell(row=row_idx, column=5).value
        gateway = sheet.cell(row=row_idx, column=7).value
        
        if not router_name or not ip_ssh: 
            continue
        network_data.append({"row_idx": row_idx, "ip": ip_ssh, "router_name": router_name, "vlan_name": vlan_name, "gateway": gateway})

    devices = {}
    for data in network_data:
        ip = data["ip"]
        if ip not in devices: 
            devices[ip] = {"name": data["router_name"], "configs": []}
        devices[ip]["configs"].append(data)

    for ip, dev_info in devices.items():
        print(f"\n[-] MEMBERSIHKAN CONFIG: {dev_info['name']} ({ip})")
        
        mikrotik_device = {
            'device_type': 'mikrotik_routeros',
            'host': ip,
            'username': SSH_USER,
            'password': SSH_PASS,
            'global_delay_factor': 2,
        }
        
        try:
            net_connect = ConnectHandler(**mikrotik_device)
            
            for config in reversed(dev_info["configs"]):
                vlan = config["vlan_name"]
                bridge_name = f"bridge-{vlan}"
                pool_name = f"pool-{vlan}"
                dhcp_name = f"dhcp-{vlan}"
                gw_ip = config["gateway"].split('/')[0]
                
                print(f"    [<] Menghapus DHCP Server & Network untuk {vlan}...")
                net_connect.send_command(f"/ip dhcp-server remove [find name={dhcp_name}]")
                net_connect.send_command(f"/ip dhcp-server network remove [find gateway={gw_ip}]")
                net_connect.send_command(f"/ip pool remove [find name={pool_name}]")
                net_connect.send_command(f"/ip address remove [find comment=\"IP-{vlan}\"]")
                
                print(f"    [<] Menghapus VLAN Interface & Mengosongkan {bridge_name}...")
                net_connect.send_command(f"/interface vlan remove [find name={vlan}]")
                net_connect.send_command(f"/interface bridge port remove [find bridge={bridge_name}]")
                net_connect.send_command(f"/interface bridge remove [find name={bridge_name}]")
                time.sleep(0.5)
            
            # Flush perubahan pasca pembersihan
            net_connect.send_command("/system backup save name=autosave_python")
            time.sleep(1)
            
            for config in dev_info["configs"]:
                sheet.cell(row=config["row_idx"], column=9).value = "Pending"
                
            print(f"[+] {dev_info['name']} bersih total.")
            net_connect.disconnect()
            
        except Exception as e:
            print(f"[!] Gagal melakukan rollback pada {dev_info['name']}: {e}")

    wb.save(EXCEL_FILE)
    print("\n[+] Status excel dikembalikan ke posisi 'Pending'.")

if __name__ == "__main__":
    main()