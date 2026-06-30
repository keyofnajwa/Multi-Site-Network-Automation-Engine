from netmiko import ConnectHandler
import openpyxl
import os
import time

EXCEL_FILE = "tabel1.xlsx"
SSH_USER = "admin"
SSH_PASS = ""

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"[!] File {EXCEL_FILE} tidak ditemukan. Pastikan file berada di direktori yang sama.")
        return

    print(f"[*] Membuka dan membaca data dari file {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    network_data = []
    for row_idx in range(4, sheet.max_row + 1):
        router_name = sheet.cell(row=row_idx, column=2).value
        ip_ssh = sheet.cell(row=row_idx, column=3).value
        vlan_id = sheet.cell(row=row_idx, column=4).value
        vlan_name = sheet.cell(row=row_idx, column=5).value
        interface = sheet.cell(row=row_idx, column=6).value
        gateway = sheet.cell(row=row_idx, column=7).value
        pool_raw = sheet.cell(row=row_idx, column=8).value
        
        if not router_name or not ip_ssh:
            continue
            
        pool_range = str(pool_raw).strip().replace(',', '-').replace(' ', '')
        
        # Ekstraksi Network Prefix dan Gateway IP
        ip_base = gateway.split('/')[0]
        prefix = gateway.split('/')[1]
        
        octets = ip_base.split('.')
        # Logika fleksibel untuk membaca network prefix baik /24 maupun /31
        if prefix == '31':
            last_octet = int(octets[3])
            network_octet = last_octet if last_octet % 2 == 0 else last_octet - 1
            network_prefix = f"{octets[0]}.{octets[1]}.{octets[2]}.{network_octet}/31"
        else:
            network_prefix = f"{octets[0]}.{octets[1]}.{octets[2]}.0/{prefix}"
            
        gateway_ip = ip_base

        network_data.append({
            "row_idx": row_idx,
            "router_name": router_name,
            "ip": ip_ssh,
            "vlan_id": int(vlan_id),
            "vlan_name": vlan_name,
            "interface": interface,
            "gateway": gateway,
            "pool_range": pool_range,
            "network": network_prefix,
            "gateway_ip": gateway_ip
        })

    devices = {}
    for data in network_data:
        ip = data["ip"]
        if ip not in devices:
            devices[ip] = {"name": data["router_name"], "configs": []}
        devices[ip]["configs"].append(data)

    for ip, dev_info in devices.items():
        print(f"\n==================== EKSEKUSI ROUTER: {dev_info['name']} ({ip}) ====================")
        
        mikrotik_device = {
            'device_type': 'mikrotik_routeros',
            'host': ip,
            'username': SSH_USER,
            'password': SSH_PASS,
            'port': 22,
            'global_delay_factor': 2,
            'timeout': 30,
        }
        
        try:
            print(f"[*] Menghubungi {dev_info['name']} via SSH...")
            net_connect = ConnectHandler(**mikrotik_device)
            success_rows = []
            
            for config in dev_info["configs"]:
                v_id = config["vlan_id"]
                v_name = config["vlan_name"]  
                iface = config["interface"]   
                gw = config["gateway"]
                p_range = config["pool_range"]
                net_add = config["network"]
                gw_ip = config["gateway_ip"]
                
                bridge_name = f"bridge-{v_name}"
                pool_name = f"pool-{v_name}"
                dhcp_name = f"dhcp-{v_name}"

                print(f"\n[>] Mengirimkan parameter konfigurasi ke {v_name}:")
                
                # 1. Membuat Bridge Induk & set PVID Bridge Induk secara eksplisit
                net_connect.send_command(f"/interface bridge add name={bridge_name} vlan-filtering=yes pvid={v_id} comment=\"Bridge {v_name}\"")
                time.sleep(0.5)
                
                # 2. Menghubungkan Port Fisik ke Bridge & Mengunci PVID Port Sesuai VLAN ID
                net_connect.send_command(f"/interface bridge port add bridge={bridge_name} interface={iface} pvid={v_id}")
                time.sleep(0.5)
                
                # 3. Membuat Sub-Interface VLAN
                net_connect.send_command(f"/interface vlan add name={v_name} vlan-id={v_id} interface={bridge_name}")
                time.sleep(0.5)
                
                # 4. Memasang IP Gateway pada Interface Bridge
                net_connect.send_command(f"/ip address add address={gw} interface={bridge_name} comment=\"IP-{v_name}\"")
                time.sleep(0.5)
                
                # ================= PROSES STANDAR DHCP SETUP  =================
                
                print(f"    -> [DHCP] Membuat Pool {pool_name}...")
                net_connect.send_command(f"/ip pool add name={pool_name} ranges={p_range}")
                time.sleep(0.5)
                
                print(f"    -> [DHCP] Menambahkan Network {net_add}...")
                net_connect.send_command(f"/ip dhcp-server network add address={net_add} gateway={gw_ip} dns-server=8.8.8.8")
                time.sleep(0.5)
                
                print(f"    -> [DHCP] Mengaktifkan Server {dhcp_name}...")
                cmd_dhcp_fix = f"/ip dhcp-server add name={dhcp_name} interface={bridge_name}address-pool={pool_name} disabled=no"
                out_dhcp = net_connect.send_command(cmd_dhcp_fix)
                
                if out_dhcp: 
                    print(f"        ❌ KESALAHAN MIKROTIK: {out_dhcp}")
                else:
                    print(f"        ✅ DHCP Server {dhcp_name} BERHASIL AKTIF.")
                time.sleep(0.5)
                
                # ==================================================================================
                # ==================================================================================
                
                success_rows.append(config["row_idx"])
            
            print(f"\n[*] Sinkronisasi Penyimpanan Permanen RAM ke Storage Perangkat...")
            net_connect.send_command("/system backup save name=autosave_python")
            time.sleep(2) 
            
            print(f"[+] Proses konfigurasi pada {dev_info['name']} sukses diterapkan.")
            for r_idx in success_rows:
                sheet.cell(row=r_idx, column=9).value = "SUCCESS"
                
            net_connect.disconnect()
            
        except Exception as e:
            print(f"[!] Gagal terhubung atau eksekusi putus pada {dev_info['name']}: {e}")
            for config in dev_info["configs"]:
                sheet.cell(row=config["row_idx"], column=9).value = "FAILED"

    wb.save(EXCEL_FILE)
    print("\n[+] Eksekusi perbaikan selesai secara keseluruhan.")

if __name__ == "__main__":
    main()